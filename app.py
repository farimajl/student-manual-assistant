from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
from flask_session import Session
from llama_index import VectorStoreIndex, ServiceContext, download_loader, Document
from llama_index.llms import OpenAI
from llama_index.embeddings import OpenAIEmbedding
from dotenv import load_dotenv
import os
import fitz  # PyMuPDF
from openai import ChatCompletion
import re
import uuid
import difflib
import unidecode
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
import openpyxl

load_dotenv()

app = Flask(__name__, template_folder="templates", static_folder="static")
CORS(app, resources={r"/chat": {"origins": "*"}}, supports_credentials=True)
app.config['SESSION_TYPE'] = 'filesystem'
app.secret_key = os.getenv("FLASK_SECRET_KEY", str(uuid.uuid4()))
Session(app)

# ==== Load and clean text from PDFs and Excel ====
doc_dir = "./documents"
def load_sentences():
    sentences = []
    for filename in os.listdir(doc_dir):
        if filename.endswith(".pdf"):
            path = os.path.join(doc_dir, filename)
            doc = fitz.open(path)
            for page in doc:
                blocks = page.get_text("blocks")
                for block in blocks:
                    if isinstance(block, tuple) and len(block) > 4:
                        text = block[4].strip()
                        for sentence in re.split(r'(?<=[.!?])\s+', text):
                            clean = sentence.strip()
                            if len(clean) > 40 or re.search(r"[0-9]{1,2}\s*EC", clean, re.IGNORECASE):
                                sentences.append(clean)
            doc.close()
    return sentences

def load_excel_sentences():
    excel_sentences = []
    global excel_documents
    excel_documents = []
    for filename in os.listdir(doc_dir):
        if filename.endswith(".xlsx"):
            try:
                path = os.path.join(doc_dir, filename)
                wb = openpyxl.load_workbook(path, data_only=True)
                for sheet in wb.worksheets:
                    print(f"📄 Processing Excel sheet: {sheet.title}")
                    header_row = [cell.value for cell in sheet[3]]
                    for r in range(5, sheet.max_row + 1):
                        row_data = {
                            'day': sheet.cell(row=r, column=2).value,
                            'hour': sheet.cell(row=r, column=3).value,
                        }
                        for c in range(4, sheet.max_column + 1):
                            cell = sheet.cell(row=r, column=c)
                            week = header_row[c - 1]
                            if cell.value:
                                context = f"Week: {week}, Day: {row_data['day']}, Hour: {row_data['hour']}, Course: {cell.value}"
                                excel_sentences.append(context)
                                excel_documents.append(Document(text=context))
            except Exception as e:
                print(f"❌ Failed to open Excel file: {filename}")
                print("   Reason:", e)
                continue
    return excel_sentences

SENTENCES = load_sentences() + load_excel_sentences()

# ==== Set up OpenAI + LlamaIndex ====
llm = OpenAI(model="gpt-3.5-turbo", api_key=os.getenv("OPENAI_API_KEY"))
embed_model = OpenAIEmbedding(api_key=os.getenv("OPENAI_API_KEY"))
service_context = ServiceContext.from_defaults(llm=llm, embed_model=embed_model)

PyMuPDFReader = download_loader("PyMuPDFReader")
loader = PyMuPDFReader()
documents = []
for filename in os.listdir(doc_dir):
    if filename.endswith(".pdf"):
        path = os.path.join(doc_dir, filename)
        documents.extend(loader.load(file_path=path))

if 'excel_documents' in globals():
    documents.extend(excel_documents)

index = VectorStoreIndex.from_documents(documents, service_context=service_context)
retriever = index.as_retriever(similarity_top_k=50)

# ==== General replies ====
GENERAL_REPLIES = {
    "hi": "Hello! I'm your assistant from the Civil Engineering Department of Twente University 😊",
    "hello": "Hi there! How can I assist you today?",
    "how are you": "I'm just a bot, but I'm ready to help you!",
    "goodbye": "Goodbye! If you have more questions later, just ask 😊",
    "bye": "Bye! See you next time.",
    "thank you": "You're welcome!",
    "thanks": "Always happy to help!",
    "ok": "Got it! Let me know if you have more questions.",
    "yes": "Alright! What would you like to know more about?",
    "no": "Okay. Let me know if anything comes up."
}

def match_general_reply(cleaned_input):
    for key in GENERAL_REPLIES:
        if cleaned_input == key:
            return GENERAL_REPLIES[key]
        if difflib.SequenceMatcher(None, cleaned_input, key).ratio() > 0.87:
            return GENERAL_REPLIES[key]
    return None

# ==== Memory ====
user_context_memory = {}

# ==== Pronoun resolution ====
def resolve_pronouns(user_input, history):
    context = "\n".join(history[-6:])
    try:
        result = ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Rewrite only the user's follow-up question using the conversation history to clarify vague references. Replace pronouns like 'his', 'her', 'this module' with the correct person or subject from history."},
                {"role": "user", "content": f"History:\n{context}\nFollow-up: {user_input}\nRewritten:"}
            ],
            api_key=os.getenv("OPENAI_API_KEY")
        )
        return result["choices"][0]["message"]["content"].strip()
    except Exception as e:
        print("Clarification error:", e)
        return user_input

# ==== TF-IDF fallback ====
def find_relevant_sentences(query: str, max_hits=30):
    if not SENTENCES:
        return ""
    vectorizer = TfidfVectorizer().fit([query] + SENTENCES)
    query_vec = vectorizer.transform([query])
    doc_vecs = vectorizer.transform(SENTENCES)
    sims = cosine_similarity(query_vec, doc_vecs).flatten()
    top = np.argsort(sims)[::-1][:max_hits]
    return "\n".join([SENTENCES[i] for i in top if sims[i] > 0.05])

def extract_matching_email_lines(clarified, context_text):
    name_tokens = set(unidecode.unidecode(clarified).lower().split())
    lines = []
    for line in context_text.split('\n'):
        lower_line = unidecode.unidecode(line).lower()
        if '@' in line:
            for token in name_tokens:
                if len(token) > 2 and token in lower_line:
                    lines.append(line)
                    break
    return lines

# ==== Chat route ====
@app.route('/chat', methods=['POST'])
def chat():
    try:
        user_input = request.json.get('message', '').strip()
        if not user_input:
            return jsonify({"response": "No input provided"}), 400

        cleaned_input = unidecode.unidecode(user_input).strip("?!.").lower()
        session_id = request.remote_addr or str(uuid.uuid4())
        user_context_memory.setdefault(session_id, []).append(user_input)

        reply = match_general_reply(cleaned_input)
        if reply:
            return jsonify({"response": reply})

        clarified = resolve_pronouns(user_input, user_context_memory[session_id])
        context_nodes = retriever.retrieve(clarified)
        node_texts = [n.get_text() for n in context_nodes if n.get_text()] if context_nodes else []

        if not node_texts or len(" ".join(node_texts)) < 50:
            node_texts += [find_relevant_sentences(clarified)]

        all_context = "\n".join(node_texts[:20]).strip()

        if "email" in clarified.lower():
            email_lines = extract_matching_email_lines(clarified, all_context)
            context_block = "\n".join(email_lines) if email_lines else all_context
        else:
            context_block = all_context

        if context_block:
            result = ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are a helpful, accurate assistant for Civil Engineering students at Twente University. Use ONLY the context below. Do NOT guess or invent. If uncertain, reply 'Nothing found'. Always consider all facts before answering."},
                    {"role": "user", "content": f"Context:\n{context_block}\n\nQuestion: {clarified}"}
                ],
                api_key=os.getenv("OPENAI_API_KEY")
            )
            reply = result["choices"][0]["message"]["content"].strip()
            return jsonify({"response": reply})

        return jsonify({"response": "Nothing found"})

    except Exception as e:
        print("Error during /chat:", e)
        return jsonify({"response": f"An error occurred: {str(e)}"}), 500

# ==== Homepage ====
@app.route('/')
def home():
    return render_template("index.html")

if __name__ == '__main__':
    app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
